import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from documents.spreadsheet.spreadsheet_document import SpreadsheetDocument
from models.spreadsheet_models import BorderProblem, SheetXmlInfo
from utils.text_utils import remove_all_spaces
from utils.xml_debug import dump_zip_structure_pretty

class ExcelDocument(SpreadsheetDocument):
    def __init__(self, path: str):
        self.path = path
        self.wb = load_workbook(path, data_only=False)
        self.wb_values = load_workbook(path, data_only=True)
        self._zip = zipfile.ZipFile(path)
        self.workbook_xml = self._load_xml("xl/workbook.xml")
        self.sheets = self._load_sheets_xml()

    STYLE_EQUIV = {
        "thick": {"thick", "medium"},
        "medium": {"medium"},
        "thin": {"thin"},
        "dotted": {"dotted"},
        "dashDot": {"dashDot"},
        "dashDotDot": {"dashDotDot"},
        "dashed": {"dashed"},
        "double": {"double"},
        "slantDashDot": {"slantDashDot"},
    }

    def _load_xml(self, name: str) -> ET.Element:
        """
        Načte XML soubor z XLSX archivu.

        Args:
            name: Vnitřní cesta k XML souboru.

        Returns:
            Kořenový XML element.
        """
        with self._zip.open(name) as f:
            return ET.fromstring(f.read())

    def _load_sheets_xml(self) -> dict[str, SheetXmlInfo]:
        """
        Načte XML reprezentace listů a vytvoří mapu podle názvu listu.

        Returns:
            Slovník ve tvaru název listu -> informace o XML listu.
        """
        sheets: dict[str, SheetXmlInfo] = {}

        rels = self._load_xml("xl/_rels/workbook.xml.rels")
        rel_map = {
            r.attrib["Id"]: r.attrib["Target"]
            for r in rels.findall(".//{*}Relationship")
        }

        for sheet in self.workbook_xml.findall(".//{*}sheet"):
            name = sheet.attrib.get("name")

            r_id = None
            for k, v in sheet.attrib.items():
                if k.endswith("}id"):
                    r_id = v
                    break

            if not r_id or not name:
                continue

            target = rel_map.get(r_id)
            if not target:
                continue

            xml = self._load_xml(f"xl/{target}")
            sheets[name] = SheetXmlInfo(
                xml=xml,
                path=target,
            )

        return sheets

    def sheet_names(self) -> list[str]:
        return self.wb.sheetnames

    def get_cell(
        self,
        address: str,
        *,
        include_value: bool = False,
    ) -> dict[str, object] | None:
        if "!" not in address:
            raise ValueError("Musí být formát 'sheet!A1'")

        sheet, addr = address.split("!", 1)

        cell = self._get_cell_obj(sheet, addr)
        if cell is None:
            return None

        if cell.value is None and cell.data_type != "f":
            return None

        data = {
            "sheet": sheet,
            "address": addr,
            "raw_cell": cell,
            "formula": cell.value if cell.data_type == "f" else None,
            "value_cached": self.wb_values[sheet][addr].value,
        }

        if include_value:
            data["value"] = cell.value

        return data

    def save_debug_xml(self, out_dir: str | Path = "debug") -> None:
        dump_zip_structure_pretty(self.path, Path(out_dir) / "excel")

    def get_array_formula_cells(self) -> list[str]:
        cells: list[str] = []

        for sheet_name, data in self.sheets.items():
            xml = data.xml

            for c in xml.iter():
                if not c.tag.endswith("}c"):
                    continue

                f = None
                for ch in list(c):
                    if ch.tag.endswith("}f"):
                        f = ch
                        break

                if f is None:
                    continue

                if f.get("t") == "array":
                    addr = c.get("r")
                    ref = f.get("ref")
                    if ref:
                        cells.append(f"{sheet_name}!{ref} (anchor {addr})")
                    else:
                        cells.append(f"{sheet_name}!{addr}")

        return cells

    def get_cell_info(self, sheet: str, addr: str) -> dict[str, object] | None:
        cell = self.get_cell(f"{sheet}!{addr}")
        if cell is None:
            return None

        formula = cell.get("formula")
        value = cell.get("value_cached")

        return {
            "exists": True,
            "formula": formula,
            "value_cached": value,
            "is_error": isinstance(value, str) and value.startswith("#"),
        }

    def _iter_formula_cells(self) -> Iterator[dict[str, str]]:
        """
        Vrací iterátor buněk obsahujících vzorec.

        Yields:
            Slovník s názvem listu, adresou buňky a vzorcem.
        """
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f" and isinstance(cell.value, str):
                        yield {
                            "sheet": ws.title,
                            "address": cell.coordinate,
                            "formula": cell.value,
                        }

    def iter_formulas(self) -> Iterator[dict[str, str]]:
        for item in self._iter_formula_cells():
            yield {"sheet": item["sheet"], "formula": item["formula"]}

    def normalize_formula(self, f: str | None) -> str:
        if f is None:
            return ""

        if hasattr(f, "text"):
            f = f.text

        if not isinstance(f, str):
            return ""

        f = remove_all_spaces(f)
        return f.upper()

    def get_defined_names(self) -> set[str]:
        return {name.upper() for name in self.wb.defined_names.keys()}

    def cells_with_formulas(self) -> list[dict[str, str]]:
        return list(self._iter_formula_cells())

    def merged_ranges(self, sheet: str) -> list[tuple[int, int, int, int]]:
        ws = self._get_sheet(sheet)
        if ws is None:
            return []
        ranges = []

        for r in ws.merged_cells.ranges:
            ranges.append(range_boundaries(str(r)))

        return ranges

    def has_conditional_formatting(self, sheet: str) -> bool:
        ws = self._get_sheet(sheet)
        if ws is None:
            return False
        return bool(ws.conditional_formatting._cf_rules)

    def check_conditional_formatting(
        self, sheet_name: str, expected: dict
    ) -> list[dict]:
        sheet = self._get_sheet(sheet_name)
        if sheet is None:
            return [{"kind": "sheet_missing", "sheet": sheet_name}]

        problems: list[dict] = []

        for cell_addr, rules in expected.items():
            actual_rules = self._get_cell_conditional_rules(sheet, cell_addr)

            for exp in rules:
                ok, reason = self._rule_exists(actual_rules, exp)
                if not ok:
                    problems.append(
                        {
                            "kind": "missing_rule",
                            "cell": cell_addr,
                            "operator": exp.get("operator"),
                            "value": exp.get("value"),
                            "reason": reason,
                        }
                    )

        return problems

    def _rule_exists(self, actual_rules, expected_rule) -> tuple[bool, str | None]:
        """
        Ověří, zda mezi pravidly existuje očekávané podmíněné formátování.

        Args:
            actual_rules: Skutečně nalezená pravidla.
            expected_rule: Očekávané pravidlo.

        Returns:
            Dvojici (nalezeno, důvod).
        """
        for rule in actual_rules:
            if rule.type != "cellIs":
                continue
            if rule.operator != expected_rule["operator"]:
                continue

            got = self._to_number(rule.formula[0] if rule.formula else None)
            want = self._to_number(expected_rule.get("value"))

            if got is None or want is None:
                continue

            if abs(got - want) < 0.01:
                return True, None

        return False, "no_match"

    def _get_cell_conditional_rules(self, sheet, addr: str) -> list:
        """
        Vrátí pravidla podmíněného formátování pro danou buňku.

        Args:
            sheet: Worksheet objekt.
            addr: Adresa buňky.

        Returns:
            Seznam pravidel.
        """
        rules = []

        for cf in sheet.conditional_formatting:
            if addr in cf.cells:
                rules.extend(cf.rules)

        return rules

    def _get_sheet(self, sheet_name: str):
        """
        Vrátí worksheet podle názvu listu.

        Args:
            sheet_name: Název listu.

        Returns:
            Worksheet, nebo None.
        """
        try:
            return self.wb[sheet_name]
        except KeyError:
            return None

    def get_cell_style(self, sheet: str, addr: str) -> dict | None:
        ws = self._get_sheet(sheet)
        if ws is None:
            return None

        cell = ws[addr]

        return {
            "number_format": cell.number_format,
            "align_h": cell.alignment.horizontal,
            "bold": bool(cell.font and cell.font.bold),
            "wrap": bool(cell.alignment and cell.alignment.wrap_text),
        }

    def iter_cells(self, sheet: str) -> Iterator[str]:
        ws = self._get_sheet(sheet)
        if ws is None:
            return
        for row in ws.iter_rows():
            for cell in row:
                yield cell.coordinate

    def _get_cell_obj(self, sheet: str, addr: str):
        """
        Vrátí raw openpyxl objekt buňky.

        Args:
            sheet: Název listu.
            addr: Adresa buňky.

        Returns:
            Objekt buňky, nebo None.
        """
        ws = self._get_sheet(sheet)
        if ws is None:
            return None
        return ws[addr]

    def get_cell_value(self, sheet: str, addr: str):
        cell = self._get_cell_obj(sheet, addr)
        return None if cell is None else cell.value

    def has_formula(self, sheet: str, addr: str) -> bool:
        cell = self._get_cell_obj(sheet, addr)
        if cell is None:
            return False
        v = cell.value
        return isinstance(v, str) and v.startswith("=")

    def has_chart(self, sheet: str) -> bool:
        return self._chart(sheet) is not None

    def _chart(self, sheet: str):
        """
        Vrátí první graf z daného listu.

        Args:
            sheet: Název listu.

        Returns:
            Objekt grafu, nebo None.
        """
        ws = self._get_sheet(sheet)
        return ws._charts[0] if ws and ws._charts else None

    def _read_rich_text(self, tx) -> str | None:
        """
        Přečte text z rich-text struktury grafu.

        Args:
            tx: Textová struktura grafu.

        Returns:
            Výsledný text, nebo None.
        """
        if not tx or not getattr(tx, "rich", None):
            return None

        parts = []
        for p in getattr(tx.rich, "p", []) or []:
            for r in getattr(p, "r", []) or []:
                t = getattr(r, "t", None)
                if t:
                    parts.append(t)

        text = "".join(parts).strip()
        return text or None

    def chart_title(self, sheet: str) -> str | None:
        c = self._chart(sheet)
        if not c or not c.title:
            return None
        return self._read_rich_text(getattr(c.title, "tx", None))

    def chart_x_label(self, sheet: str) -> str | None:
        c = self._chart(sheet)
        if c and c.x_axis and c.x_axis.title:
            return self._read_rich_text(getattr(c.x_axis.title, "tx", None))
        return None

    def chart_y_label(self, sheet: str) -> str | None:
        c = self._chart(sheet)
        if c and c.y_axis and c.y_axis.title:
            return self._read_rich_text(getattr(c.y_axis.title, "tx", None))
        return None

    def chart_has_data_labels(self, sheet: str) -> bool:
        c = self._chart(sheet)
        if not c:
            return False
        return any(s.dLbls and s.dLbls.showVal for s in c.series)

    def chart_type(self, sheet: str) -> str | None:
        c = self._chart(sheet)
        if not c:
            return None

        tag = (c.tagname or "").lower()

        if "bar" in tag:
            return "bar"
        if "line" in tag:
            return "line"
        if "pie" in tag:
            return "pie"

        return tag

    def has_3d_chart(self, sheet: str) -> bool:
        ws = self._get_sheet(sheet)
        if not ws:
            return False

        for chart in ws._charts:
            tag = chart.tagname.lower()
            if "3d" in tag:
                return True

        return False

    def has_sheet(self, name: str) -> bool:
        return self._get_sheet(name) is not None

    def _matches(self, actual, expected):
        """
        Ověří shodu skutečného a očekávaného stylu čáry.

        Args:
            actual: Skutečný styl.
            expected: Očekávaný styl.

        Returns:
            True, pokud styly odpovídají.
        """
        allowed = self.STYLE_EQUIV.get(expected, {expected})
        return actual in allowed

    def check_table_borders(
        self,
        sheet: str,
        location: str,
        outer: str,
        inner: str,
    ) -> list[BorderProblem]:
        ws = self._get_sheet(sheet)
        if ws is None:
            return [BorderProblem(kind="sheet_missing", sheet=sheet)]

        problems: list[BorderProblem] = []
        min_col, min_row, max_col, max_row = range_boundaries(location)

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                b = cell.border

                if r == min_row and not self._matches(b.top.style, outer):
                    problems.append(
                        BorderProblem(
                            kind="outer",
                            side="top",
                            cell=cell.coordinate,
                            expected=outer,
                        )
                    )

                if r == max_row and not self._matches(b.bottom.style, outer):
                    problems.append(
                        BorderProblem(
                            kind="outer",
                            side="bottom",
                            cell=cell.coordinate,
                            expected=outer,
                        )
                    )

                if c == min_col and not self._matches(b.left.style, outer):
                    problems.append(
                        BorderProblem(
                            kind="outer",
                            side="left",
                            cell=cell.coordinate,
                            expected=outer,
                        )
                    )

                if c == max_col and not self._matches(b.right.style, outer):
                    problems.append(
                        BorderProblem(
                            kind="outer",
                            side="right",
                            cell=cell.coordinate,
                            expected=outer,
                        )
                    )

                if r > min_row:
                    ok = self._matches(b.top.style, inner)
                    if not ok:
                        up = ws.cell(row=r - 1, column=c).border
                        ok = self._matches(up.bottom.style, inner)
                    if not ok:
                        problems.append(
                            BorderProblem(
                                kind="inner_h",
                                cell=cell.coordinate,
                                expected=inner,
                            )
                        )

                if c > min_col:
                    ok = self._matches(b.left.style, inner)
                    if not ok:
                        left = ws.cell(row=r, column=c - 1).border
                        ok = self._matches(left.right.style, inner)
                    if not ok:
                        problems.append(
                            BorderProblem(
                                kind="inner_v",
                                cell=cell.coordinate,
                                expected=inner,
                            )
                        )

        return problems
